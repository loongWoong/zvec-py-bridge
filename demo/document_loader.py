"""文档加载器 — 支持 Markdown、DOCX、PDF、XLSX、图片的文档解析与切分。

每种格式有独立的解析函数，按文件扩展名分发。解析库延迟导入，
缺库时返回清晰错误而非崩溃。纯文本文件作为 fallback。

OCR 支持：扫描件 PDF 和图片通过 Ollama glm-ocr 模型识别为文本，
DOCX 内嵌图片也会提取并 OCR。OCR 结果转为 Markdown 存入知识库。
"""
from __future__ import annotations

import io
import re

import kb_data as kb


# ====================================================================== #
#  OCR — 调用 Ollama glm-ocr 模型识别图片文字
# ====================================================================== #
_OCR_PROMPT = (
    "请识别图片中的全部文字内容，使用Markdown格式输出。要求：\n"
    "1. 表格必须使用Markdown表格语法（| 列1 | 列2 |）保留原始表格结构\n"
    "2. 标题使用 # 标记\n"
    "3. 列表使用 - 或 1. 标记\n"
    "4. 不要输出HTML标签\n"
    "5. 不要重复输出相同内容\n"
)


def ocr_image(image_bytes: bytes) -> str:
    """调用 Ollama OCR 模型识别图片中的文字，返回 Markdown 格式文本。

    大图（高度 > 1000px 或宽度 > 1600px）会被拆分为多段分别 OCR，
    每段独立去重后再合并，避免模型因缩放导致部分文字无法识别。
    """
    import io
    from PIL import Image

    img = Image.open(io.BytesIO(image_bytes))
    width, height = img.size
    print(f"\n  [OCR] 图片尺寸: {width}x{height}")

    # 宽度过大时先等比缩小（避免 Ollama 缩放过多导致文字模糊）
    if width > 1600:
        new_width = 1600
        new_height = int(height * (new_width / width))
        img = img.resize((new_width, new_height), Image.LANCZOS)
        width, height = img.size
        print(f"  [OCR] 宽度缩放: {width}x{height}")

    # 判断是否需要按高度拆分（阈值较低，确保大图都能被分段处理）
    need_split = height > 800

    if not need_split:
        # 小图：直接 OCR → 后处理
        buf = io.BytesIO()
        img.save(buf, format="PNG")
        raw = _ocr_single(buf.getvalue())
        text = _postprocess_ocr(raw)
        print(f"  [OCR] 单张识别: {len(text)} chars")
    else:
        # 大图：上下拆分 → 每段 OCR + 后处理 → 合并
        chunk_h = 700
        overlap = 100
        parts: list[str] = []
        y = 0
        part_num = 0
        while y < height:
            bottom = min(y + chunk_h, height)
            crop = img.crop((0, y, width, bottom))
            buf = io.BytesIO()
            crop.save(buf, format="PNG")
            part_num += 1
            print(f"  [OCR] 拆分段 {part_num}: y={y}-{bottom} ({bottom - y}px)")
            raw = _ocr_single(buf.getvalue())
            cleaned = _postprocess_ocr(raw)
            if cleaned:
                parts.append(cleaned)
            y = bottom - overlap if bottom < height else height

        text = _combine_chunks(parts)
        print(f"  [OCR] 合并 {len(parts)} 段: {len(text)} chars")

    print(f"{'='*60}")
    print(text[:4000])
    if len(text) > 4000:
        print(f"  ... (总长 {len(text)} chars)")
    print(f"{'='*60}\n")
    return text


def _postprocess_ocr(raw: str) -> str:
    """对单段 OCR 原始输出做后处理：表格格式化 + HTML转换 + 去重。"""
    if not raw:
        return ""
    text = format_tabular_to_markdown(raw)
    text = html_tables_to_markdown(text)
    text = deduplicate_ocr(text)
    return text


def _ocr_single(image_bytes: bytes) -> str:
    """对单张图片调用 OCR 模型，返回原始文本。"""
    import base64
    import requests

    img_b64 = base64.b64encode(image_bytes).decode()
    r = requests.post(f"{kb.OCR_URL}/api/generate", json={
        "model": kb.OCR_MODEL,
        "prompt": _OCR_PROMPT,
        "images": [img_b64],
        "stream": False,
        "options": {"temperature": 0.1, "num_predict": 4000},
    }, timeout=180)
    if r.status_code != 200:
        print(f"  [OCR] 错误: HTTP {r.status_code} - {r.text[:200]}")
        return ""
    return r.json().get("response", "").strip()


def _combine_chunks(parts: list[str]) -> str:
    """合并多段 OCR 结果，去除重叠区域的重复行。"""
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0]

    combined_lines = parts[0].split("\n")
    for i in range(1, len(parts)):
        next_lines = [l.strip() for l in parts[i].split("\n") if l.strip()]
        if not next_lines:
            continue
        # 查找重叠行数：combined 末尾与 next 开头相同的行
        overlap = 0
        for j in range(min(len(combined_lines), len(next_lines), 15), 0, -1):
            tail = [l.strip() for l in combined_lines[-j:] if l.strip()]
            head = next_lines[:j]
            if tail == head:
                overlap = j
                break
        combined_lines.extend(next_lines[overlap:])

    return "\n".join(combined_lines)


# ====================================================================== #
#  表格检测与格式化 — 将空格/制表符对齐的表格数据转为 Markdown 表格
# ====================================================================== #
def format_tabular_to_markdown(text: str) -> str:
    """检测 OCR 输出中的表格数据（制表符/多空格分隔），转为 Markdown 表格。

    OCR 模型经常输出用空格对齐的表格数据，但不使用 Markdown 语法。
    此函数检测连续的多列行，自动转换为 Markdown 表格格式。
    """
    lines = text.split("\n")
    result: list[str] = []
    i = 0

    while i < len(lines):
        cols = _split_columns(lines[i])

        if cols and len(cols) >= 2:
            col_count = len(cols)
            table_rows = [cols]
            j = i + 1
            while j < len(lines):
                row_cols = _split_columns(lines[j])
                if not row_cols:
                    break
                # 尝试调整列数以匹配表头
                row_cols = _adjust_columns(row_cols, col_count)
                table_rows.append(row_cols)
                j += 1

            if len(table_rows) >= 2:
                result.append(_to_markdown_table(table_rows, col_count))
                i = j
                continue

        result.append(lines[i])
        i += 1

    return "\n".join(result)


def _split_columns(line: str) -> list[str] | None:
    """将一行文本按制表符或 2+ 空格分割为列。"""
    line = line.rstrip()
    if not line.strip():
        return None

    # 优先按制表符分割
    if "\t" in line:
        cols = [c.strip() for c in line.split("\t")]
        cols = [c for c in cols if c]
        return cols if len(cols) >= 2 else None

    # 按 2+ 空格分割
    parts = re.split(r" {2,}", line.strip())
    cols = [p.strip() for p in parts if p.strip()]
    return cols if len(cols) >= 2 else None


def _adjust_columns(cols: list[str], expected: int) -> list[str]:
    """调整列数以匹配期望值。

    列数不足时：尝试将最后一列按短标记（如 ✓ ✗ 是 否）拆分为两列。
    列数过多时：合并多余的列到最后一列。
    """
    if len(cols) == expected:
        return cols

    if len(cols) < expected:
        # 保留除最后一列外的所有列，只尝试拆分最后一列
        result = list(cols[:-1])
        last_col = cols[-1]
        if len(result) < expected - 1:
            # 尝试拆分：短标记(1-3字符) + 空格 + 剩余内容
            m = re.match(r"^(\S{1,3})\s+(.+)", last_col)
            if m and len(result) + 2 <= expected:
                result.append(m.group(1))
                result.append(m.group(2))
            else:
                result.append(last_col)
        else:
            result.append(last_col)
        while len(result) < expected:
            result.append("")
        return result[:expected]

    # 列数过多，合并
    return cols[: expected - 1] + [" ".join(cols[expected - 1:])]


def _to_markdown_table(rows: list[list[str]], col_count: int) -> str:
    """将行数据转为 Markdown 表格字符串。"""
    md: list[str] = []
    # 表头
    header = rows[0][:col_count]
    while len(header) < col_count:
        header.append("")
    md.append("| " + " | ".join(c.strip() for c in header) + " |")
    md.append("| " + " | ".join(["---"] * col_count) + " |")
    # 数据行
    for row in rows[1:]:
        while len(row) < col_count:
            row.append("")
        md.append("| " + " | ".join(c.strip() for c in row[:col_count]) + " |")
    return "\n".join(md)


def html_tables_to_markdown(text: str) -> str:
    """将模型输出的 HTML 表格转换为 Markdown 表格。"""
    def _convert_table(match: re.Match) -> str:
        html = match.group(0)
        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", html, re.DOTALL | re.IGNORECASE)
        if not rows:
            return html
        md_rows: list[str] = []
        for i, row in enumerate(rows):
            cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, re.DOTALL | re.IGNORECASE)
            cells = [re.sub(r"<[^>]+>", "", c).strip() for c in cells]
            if not cells:
                continue
            md_rows.append("| " + " | ".join(cells) + " |")
            if i == 0:
                md_rows.append("| " + " | ".join(["---"] * len(cells)) + " |")
        return "\n".join(md_rows) if md_rows else html

    return re.sub(
        r"<table[^>]*>.*?</table>", _convert_table, text, flags=re.DOTALL | re.IGNORECASE
    )


def deduplicate_ocr(text: str) -> str:
    """去除 OCR 模型重复输出的内容。

    glm-ocr 等小模型容易重复同一段文本。策略：
    1. 去掉 markdown 代码围栏
    2. 逐行尝试作为锚点，找到第一个在文本中出现两次的行
    3. 截取第二次出现之前的内容（即首次完整输出）
    """
    if not text:
        return ""
    # 去掉 markdown 代码围栏 (```markdown, ```, 等)
    text = re.sub(r"```+\w*", "", text).strip()
    if not text:
        return ""

    # 逐行尝试作为锚点，找到第一个出现两次的行
    for line in text.split("\n"):
        line = line.strip()
        if len(line) < 3:
            continue  # 跳过太短的行
        first_pos = text.find(line)
        second_pos = text.find(line, first_pos + len(line))
        if second_pos > first_pos:
            # 找到重复，截取第二次出现之前的内容
            return text[:second_pos].rstrip()

    # 没有找到重复，按空行分块去重（fallback）
    blocks = re.split(r"\n{2,}", text)
    seen: set[str] = set()
    unique_blocks: list[str] = []
    for block in blocks:
        block = block.strip()
        if not block:
            continue
        if block in seen:
            break
        seen.add(block)
        unique_blocks.append(block)
    return "\n\n".join(unique_blocks) if unique_blocks else text.strip()


# ====================================================================== #
#  主入口
# ====================================================================== #
def parse_file(filename: str, content: bytes) -> dict:
    """解析上传的文件，返回 {title, content, chunks, source_type, source_file}。

    chunks: [{heading, content}, ...]
    """
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    title = filename.rsplit(".", 1)[0] if "." in filename else filename

    if ext == "md":
        text = content.decode("utf-8", errors="replace")
        chunks = parse_markdown(text)
        source_type = "markdown"
    elif ext == "docx":
        chunks = parse_docx(content)
        source_type = "docx"
    elif ext == "pdf":
        chunks = parse_pdf(content)
        source_type = "pdf"
    elif ext in ("xlsx", "xls"):
        chunks = parse_xlsx(content)
        source_type = "xlsx"
    elif ext in ("png", "jpg", "jpeg", "bmp", "gif", "tiff", "webp"):
        chunks = parse_image(content)
        source_type = "image"
    else:
        # 纯文本 fallback
        text = content.decode("utf-8", errors="replace")
        chunks = [{"heading": "", "content": text.strip()}]
        source_type = "text"

    full_content = "\n\n".join(c["content"] for c in chunks)
    return {
        "title": title,
        "content": full_content,
        "chunks": chunks,
        "source_type": source_type,
        "source_file": filename,
    }


# ====================================================================== #
#  Markdown
# ====================================================================== #
def parse_markdown(text: str) -> list[dict]:
    """按标题切分 Markdown，保留 heading；长段进一步按句号拆分。"""
    chunks: list[dict] = []
    current_heading = ""
    current_lines: list[str] = []

    for line in text.split("\n"):
        if re.match(r"^#{1,6}\s", line):
            if current_lines:
                content = "\n".join(current_lines).strip()
                if content:
                    chunks.append({"heading": current_heading, "content": content})
            current_heading = re.sub(r"^#{1,6}\s*", "", line).strip()
            current_lines = []
        else:
            current_lines.append(line)

    if current_lines:
        content = "\n".join(current_lines).strip()
        if content:
            chunks.append({"heading": current_heading, "content": content})

    if not chunks:
        chunks = [{"heading": "", "content": text.strip()}]

    return _split_long_chunks(chunks)


# ====================================================================== #
#  DOCX — 段落 + 内嵌图片 OCR
# ====================================================================== #
def parse_docx(content: bytes) -> list[dict]:
    """用 python-docx 提取段落，识别 Heading 样式；内嵌图片走 OCR。"""
    try:
        from docx import Document as DocxDocument
    except ImportError:
        return [{"heading": "", "content": "(缺少 python-docx 依赖，无法解析 DOCX)"}]

    doc = DocxDocument(io.BytesIO(content))
    chunks: list[dict] = []
    current_heading = ""
    current_lines: list[str] = []

    for para in doc.paragraphs:
        text = para.text.strip()
        if not text:
            continue
        style_name = para.style.name if para.style else ""
        if style_name.startswith("Heading"):
            if current_lines:
                chunks.append({"heading": current_heading, "content": "\n".join(current_lines)})
            current_heading = text
            current_lines = []
        else:
            current_lines.append(text)

    if current_lines:
        chunks.append({"heading": current_heading, "content": "\n".join(current_lines)})

    # 提取内嵌图片并 OCR
    try:
        for rel in doc.part.rels.values():
            if "image" in rel.reltype:
                image_bytes = rel.target_part.blob
                ocr_text = ocr_image(image_bytes)
                if ocr_text:
                    chunks.append({"heading": "图片(OCR)", "content": ocr_text})
    except Exception:
        pass  # 图片提取失败不影响文本解析

    if not chunks:
        chunks = [{"heading": "", "content": "(空文档)"}]

    return _split_long_chunks(chunks)


# ====================================================================== #
#  PDF — 文本提取 + 扫描件 OCR 回退
# ====================================================================== #
def parse_pdf(content: bytes) -> list[dict]:
    """用 pypdf 提取文本；页面无文本时用 PyMuPDF 渲染为图片走 OCR。"""
    chunks: list[dict] = []

    # 优先用 PyMuPDF（fitz），它既能提取文本又能渲染图片
    try:
        import fitz  # PyMuPDF

        doc = fitz.open(stream=content, filetype="pdf")
        for i, page in enumerate(doc):
            text = page.get_text().strip()
            if len(text) >= 10:
                # 有足够文本，直接使用
                chunks.append({"heading": f"第{i + 1}页", "content": text})
            else:
                # 文本不足，可能是扫描件 — 渲染页面为图片走 OCR
                pix = page.get_pixmap(dpi=200)
                image_bytes = pix.tobytes("png")
                ocr_text = ocr_image(image_bytes)
                if ocr_text:
                    chunks.append({"heading": f"第{i + 1}页(OCR)", "content": ocr_text})
                else:
                    chunks.append({"heading": f"第{i + 1}页", "content": "(无法提取文本)"})
        doc.close()
    except ImportError:
        # PyMuPDF 不可用，回退到 pypdf（无 OCR 能力）
        chunks = _parse_pdf_pypdf(content)

    if not chunks:
        chunks = [{"heading": "", "content": "(无法解析 PDF)"}]

    return _split_long_chunks(chunks)


def _parse_pdf_pypdf(content: bytes) -> list[dict]:
    """pypdf 纯文本提取（无 OCR 回退）。"""
    try:
        from pypdf import PdfReader
    except ImportError:
        return [{"heading": "", "content": "(缺少 pypdf 依赖，无法解析 PDF)"}]

    reader = PdfReader(io.BytesIO(content))
    chunks: list[dict] = []
    for i, page in enumerate(reader.pages):
        text = page.extract_text()
        if text and text.strip():
            chunks.append({"heading": f"第{i + 1}页", "content": text.strip()})
    if not chunks:
        chunks = [{"heading": "", "content": "(无法提取文本，可能是扫描件，需安装 PyMuPDF 启用 OCR)"}]
    return chunks


# ====================================================================== #
#  XLSX
# ====================================================================== #
def parse_xlsx(content: bytes) -> list[dict]:
    """用 openpyxl 逐工作表提取，行用 | 连接。"""
    try:
        from openpyxl import load_workbook
    except ImportError:
        return [{"heading": "", "content": "(缺少 openpyxl 依赖，无法解析 XLSX)"}]

    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    chunks: list[dict] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        rows: list[str] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                rows.append(" | ".join(cells))
        if rows:
            content_text = "\n".join(rows)
            chunks.append({"heading": f"工作表: {sheet_name}", "content": content_text})

    if not chunks:
        chunks = [{"heading": "", "content": "(空表格)"}]

    return _split_long_chunks(chunks)


# ====================================================================== #
#  图片 — 直接 OCR
# ====================================================================== #
def parse_image(content: bytes) -> list[dict]:
    """对图片直接执行 OCR，返回识别结果。"""
    ocr_text = ocr_image(content)
    if ocr_text:
        return [{"heading": "OCR识别结果", "content": ocr_text}]
    return [{"heading": "", "content": "(无法识别图片内容)"}]


# ====================================================================== #
#  通用切分
# ====================================================================== #
def split_long_chunk(chunk: dict, max_chars: int = 500) -> list[dict]:
    """将过长的 chunk 按句号/换行切分为更小的片段。"""
    content = chunk["content"]
    heading = chunk["heading"]
    sentences = re.split(r"(?<=[。！？\n.!?])", content)
    result: list[dict] = []
    current = ""
    for sent in sentences:
        if len(current) + len(sent) > max_chars and current:
            result.append({"heading": heading, "content": current.strip()})
            current = sent
        else:
            current += sent
    if current.strip():
        result.append({"heading": heading, "content": current.strip()})
    return result if result else [chunk]


def _split_long_chunks(chunks: list[dict], max_chars: int = 500) -> list[dict]:
    """对 chunk 列表中过长的项进行二次切分。"""
    result: list[dict] = []
    for chunk in chunks:
        if len(chunk["content"]) > max_chars:
            result.extend(split_long_chunk(chunk, max_chars))
        else:
            result.append(chunk)
    return result
